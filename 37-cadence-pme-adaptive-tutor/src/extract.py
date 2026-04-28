"""Document extraction utilities for CADENCE.

Reads .docx assignment instructions and student submissions to feed the
adaptive analysis prompt. Falls back to .txt sibling if python-docx is not
installed (the synthetic generator drops a .txt twin in that case).
"""
from __future__ import annotations

from pathlib import Path


def read_docx(path: Path) -> str:
    """Return the concatenated paragraph text of a .docx."""
    try:
        from docx import Document
    except Exception:
        twin = path.with_suffix(".txt")
        if twin.exists():
            return twin.read_text()
        return ""
    if not path.exists():
        twin = path.with_suffix(".txt")
        if twin.exists():
            return twin.read_text()
        return ""
    doc = Document(str(path))
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def read_xlsx_rubric(path: Path) -> list[dict]:
    """Return a list of {axis, descriptors[], weight} rows."""
    try:
        from openpyxl import load_workbook
    except Exception:
        return []
    if not path.exists():
        return []
    wb = load_workbook(str(path))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    out = []
    header = rows[0]
    for r in rows[1:]:
        if r[0] is None:
            continue
        out.append({
            "axis": r[0],
            "descriptors": [str(x) for x in r[1:-1] if x is not None],
            "weight": r[-1] if r[-1] is not None else 1.0,
        })
    return out
